def make_chart(self, e_var, tmp_V, tmp_I, tmp_VGE, tmp_T):
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.drawing.line import LineProperties
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.layout import Layout, ManualLayout
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = Workbook()
    ws = wb.active
    df = pd.concat([tmp_V, tmp_I, tmp_VGE, tmp_T], axis=1, sort=False)
    sc = ScatterChart()
    sc2 = ScatterChart()
    maxrows = 5001
    minrows = 1
    maxcol = 5
    mincol = 2
    sc.width = 17
    sc.height = 13
    sc.y_axis.axId = 200
    sc.y_axis.scaling.min = -200
    sc.y_axis.scaling.max = 1000
    sc.x_axis.scaling.max = 5e-6
    sc2.width = 17
    sc2.height = 12
    sc2.y_axis.scaling.min = -100
    sc2.y_axis.scaling.max = 20
    sc2.x_axis.scaling.max = 5e-6
    sc2.x_axis.title = 'Time[sec]'  # X軸のタイトル
    sc.y_axis.title = 'VCE[V]'  # Y軸のタイトル
    sc2.y_axis.title = 'VGE[V]'
    # sc.x_axis.ChartLines.minorGridlines = GraphicalProperties(
    #     ln=LineProperties(solidFill="FF0000", w=0.01 * 12700, prstDash="dot"))
    # sc2.y_axis.ChartLines.minorGridlines = GraphicalProperties(
    #     ln=LineProperties(solidFill="FF0000", w=0.01 * 12700, prstDash="dot"))
    sc.legend.position = 't'
    sc2.legend.position = 't'
    sc2.x_axis.crossesAt = sc2.y_axis.scaling.min
    sc.y_axis.majorUnit = 200
    sc.x_axis.majorUnit = 1e-6
    sc.y_axis.minorUnit = 40.0
    sc.x_axis.minorUnit = 2.5e-7
    sc2.y_axis.majorUnit = 20
    sc2.y_axis.majorTickMark = "out"
    sc2.x_axis.majorUnit = 1e-6
    sc2.y_axis.minorUnit = 4.0
    sc2.x_axis.minorUnit = 2.5e-7

    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)
    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'
    ws.delete_rows(2)
    xvalues = Reference(ws, 5, minrows + 1, 5, maxrows)
    for i in range(mincol, maxcol):
        if ws.cell(1, i).value == "VGE[V]":
            yvalues2 = Reference(ws, i, minrows, maxcol - 1, maxrows)
            series2 = Series(yvalues2, xvalues, title_from_data=True)
            sc2.series.append(series2)
            sc2.y_axis.crosses = "max"
            sc2 += sc
            continue

        yvalues = Reference(ws, i, minrows, i, maxrows)
        series = Series(yvalues, xvalues, title_from_data=True)
        sc.series.append(series)

    ws.add_chart(sc2, "D6")
    wb.save(e_var + "\\pandas_openpyxl.xlsx")
